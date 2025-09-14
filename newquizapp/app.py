from flask import Flask, render_template, request, redirect, url_for, session, flash
import pymysql
pymysql.install_as_MySQLdb()
from flask_mysqldb import MySQL
import random
import string
from functools import wraps
import pandas as pd
from flask import Flask

app = Flask(__name__)


app.secret_key = 'your_secret_key_here'

# MySQL Configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'alia123'
app.config['MYSQL_DB'] = 'quizapp'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

mysql = MySQL(app)

# Login required decorator
def login_required(role=None):
    def wrapper(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'logged_in' not in session:
                flash('Please log in first.', 'danger')
                return redirect(url_for('login'))
            if role and session.get('role') != role:
                flash('Access denied.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return wrapper

# Helper to generate course code
def generate_course_code(length=6):
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/role/<role>', methods=['GET', 'POST'])
def role_register(role):
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']

        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM users WHERE email = %s", (email,))
        account = cur.fetchone()
        if account:
            flash('Email already exists.', 'danger')
        else:
            cur.execute("INSERT INTO users (username, email, password, role) VALUES (%s, %s, %s, %s)",
                        (name, email, password, role))
            mysql.connection.commit()
            flash('Registration successful. Please log in.', 'success')
            return redirect(url_for('login'))

    return render_template('register.html', role=role)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM users WHERE email = %s AND password = %s", (email, password))
        user = cur.fetchone()
        if user:
            session['logged_in'] = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials.', 'danger')

    return render_template('login.html')

@app.route('/dashboard')
@login_required()
def dashboard():
    if session['role'] == 'teacher':
        cur = mysql.connection.cursor()
        user_id = session['user_id']
        cur.execute("SELECT * FROM courses WHERE teacher_id = %s", (user_id,))
        courses = cur.fetchall()
        for course in courses:
            cur.execute("SELECT * FROM quizzes WHERE course_id = %s", (course['id'],))
            course['quizzes'] = cur.fetchall()
        return render_template('teacher_dashboard.html', courses=courses)

    else:
        return student_dashboard()  # Use the correct function that loads quizzes too


@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('home'))


@app.route('/create_course', methods=['GET', 'POST'])
@login_required('teacher')
def create_course():
    if request.method == 'POST':
        course_name = request.form['name']
        course_code = request.form['code']  # Use the code that was shown in the form
        teacher_id = session['user_id']

        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO courses (name, code, teacher_id) VALUES (%s, %s, %s)",
                    (course_name, course_code, teacher_id))
        mysql.connection.commit()
        flash(f'Course "{course_name}" created with code: {course_code}', 'success')
        return redirect(url_for('dashboard'))

    # Generate a code only once when showing the form
    course_code = generate_course_code()
    return render_template('create_course.html', code=course_code)
@app.route('/enroll', methods=['GET', 'POST'])
@login_required('student')
def enroll():
    if request.method == 'POST':
        course_code = request.form['course_code']
        student_id = session['user_id']

        cur = mysql.connection.cursor()
        # Check if course exists
        cur.execute("SELECT id FROM courses WHERE code = %s", (course_code,))
        course = cur.fetchone()

        if not course:
            flash('Invalid course code.', 'danger')
        else:
            course_id = course['id']
            # Check if already enrolled
            cur.execute("SELECT * FROM enrollments WHERE student_id = %s AND course_id = %s", (student_id, course_id))
            enrolled = cur.fetchone()
            if enrolled:
                flash('You are already enrolled in this course.', 'info')
            else:
                cur.execute("INSERT INTO enrollments (student_id, course_id) VALUES (%s, %s)", (student_id, course_id))
                mysql.connection.commit()
                flash('Enrollment successful!', 'success')
                return redirect(url_for('dashboard'))

    return render_template('enroll.html')


@app.route('/upload_quiz_excel/<int:course_id>', methods=['GET', 'POST'])
@login_required('teacher')
def upload_quiz_excel(course_id):
    if request.method == 'POST':
        topic = request.form['topic']
        file = request.files['excel_file']

        if file and file.filename.endswith('.xlsx'):
            df = pd.read_excel(file)

            expected_columns = {'Question', 'A', 'B', 'C', 'D', 'Correct'}
            if not expected_columns.issubset(df.columns):
                flash('Excel format incorrect. Required columns: Question, A, B, C, D, Correct', 'danger')
                return redirect(request.url)

            cur = mysql.connection.cursor()
            cur.execute("INSERT INTO quizzes (course_id, topic) VALUES (%s, %s)", (course_id, topic))
            quiz_id = cur.lastrowid

            for _, row in df.iterrows():
                cur.execute("""
                    INSERT INTO questions
                    (quiz_id, question, option_a, option_b, option_c, option_d, correct_option)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    quiz_id,
                    row['Question'],
                    row['A'],
                    row['B'],
                    row['C'],
                    row['D'],
                    row['Correct'].strip().upper()
                ))

            mysql.connection.commit()
            flash('Quiz uploaded successfully from Excel!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Please upload a valid .xlsx Excel file.', 'danger')

    return render_template('upload_quiz_excel.html', course_id=course_id)
from openpyxl import load_workbook
import os

from flask import request, render_template, redirect, url_for, flash
from openpyxl import load_workbook  # Make sure openpyxl is installed (pip install openpyxl)
import os

@app.route('/create_quiz/<int:course_id>', methods=['GET', 'POST'])
@login_required('teacher')
def create_quiz(course_id):
    if request.method == 'POST':
        topic = request.form['topic']
        excel_file = request.files.get('excel_file')

        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO quizzes (course_id, topic) VALUES (%s, %s)", (course_id, topic))
        quiz_id = cur.lastrowid

        # Handle Excel Upload
        if excel_file and excel_file.filename.endswith('.xlsx'):
            # Save the uploaded file temporarily
            file_path = os.path.join('uploads', excel_file.filename)
            excel_file.save(file_path)

            wb = load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is header
                question, a, b, c, d, correct = row
                cur.execute("""
                    INSERT INTO questions (quiz_id, question, option_a, option_b, option_c, option_d, correct_option)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (quiz_id, question, a, b, c, d, correct.upper()))

            # Clean up by deleting the temporary uploaded file
            os.remove(file_path)

        else:
            # Manual entry fallback
            questions = request.form.getlist('question')
            options_a = request.form.getlist('option_a')
            options_b = request.form.getlist('option_b')
            options_c = request.form.getlist('option_c')
            options_d = request.form.getlist('option_d')
            correct_options = request.form.getlist('correct_option')

            for i in range(len(questions)):
                cur.execute("""
                    INSERT INTO questions (quiz_id, question, option_a, option_b, option_c, option_d, correct_option)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    quiz_id,
                    questions[i],
                    options_a[i],
                    options_b[i],
                    options_c[i],
                    options_d[i],
                    correct_options[i].upper()
                ))

        mysql.connection.commit()
        flash('Quiz created successfully!', 'success')
        return redirect(url_for('dashboard'))

    return render_template('create_quiz.html', course_id=course_id)
@app.route('/teacher_dashboard')
@login_required('teacher')
def teacher_dashboard():
    cur = mysql.connection.cursor()

    # Get courses by teacher
    cur.execute("SELECT * FROM courses WHERE teacher_id = %s", (session['user_id'],))
    courses = cur.fetchall()

    # For each course, get its quizzes
    for course in courses:
        cur.execute("SELECT * FROM quizzes WHERE course_id = %s", (course['id'],))
        quizzes = cur.fetchall()
        course['quizzes'] = quizzes  # Attach quizzes to the course

    return render_template('teacher_dashboard.html', courses=courses)


@app.route('/student_dashboard')
@login_required('student')
def student_dashboard():
    print("Organized courses data:")
    cur = mysql.connection.cursor()

    # Debug: Print current user and session
    print(f"Current user_id: {session['user_id']}")
    print(f"Session data: {session}")

    # Get all enrolled courses with their quizzes in a single query
    cur.execute("""
        SELECT 
            c.id AS course_id,
            c.name AS course_name,
            c.code AS course_code,
            q.id AS quiz_id,
            q.topic AS quiz_topic
        FROM enrollments e
        JOIN courses c ON e.course_id = c.id
        LEFT JOIN quizzes q ON q.course_id = c.id
        WHERE e.student_id = %s
        ORDER BY c.name, q.topic
    """, (session['user_id'],))

    results = cur.fetchall()
    # Get quiz IDs the student has already attempted
    cur.execute("""
        SELECT quiz_id FROM results
        WHERE student_id = %s
    """, (session['user_id'],))
    attempted_quizzes = [row['quiz_id'] for row in cur.fetchall()]

    # Debug: Print raw query results
    print("Raw query results:", results)

    # Organize data by course
    courses = {}
    for row in results:
        course_id = row['course_id']
        if course_id not in courses:
            courses[course_id] = {
                'id': course_id,
                'name': row['course_name'],
                'code': row['course_code'],
                'quizzes': []
            }
        if row['quiz_id']:  # Only add if quiz exists
            courses[course_id]['quizzes'].append({
                'id': row['quiz_id'],
                'topic': row['quiz_topic']
            })

    # Debug: Print organized data
    print("Organized courses data:", courses)

    return render_template('student_dashboard.html', courses=list(courses.values()),
                           attempted_quiz_ids=attempted_quizzes)

@app.route('/view_results')
@login_required('student')
def view_results():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT q.id as quiz_id, q.topic, r.score, r.total, r.taken_at
        FROM results r
        JOIN quizzes q ON r.quiz_id = q.id
        WHERE r.student_id = %s
        ORDER BY r.taken_at DESC
    """, (session['user_id'],))
    results = cur.fetchall()
    return render_template('student_results.html', results=results)
@app.route('/take_quiz/<int:quiz_id>', methods=['GET', 'POST'])
@login_required()
def take_quiz(quiz_id):
    cursor = mysql.connection.cursor()  # Regular cursor (not DictCursor)
    student_id = session['user_id']

    # Check if quiz exists
    cursor.execute("SELECT * FROM quizzes WHERE id = %s", (quiz_id,))
    quiz = cursor.fetchone()

    if not quiz:
        cursor.close()
        flash("Quiz not found", "error")
        return redirect(url_for('student_dashboard'))

    # Check if already taken
    cursor.execute("SELECT * FROM results WHERE quiz_id = %s AND student_id = %s", (quiz_id, student_id))
    if cursor.fetchone():
        cursor.close()
        flash("You have already attempted this quiz.", "warning")
        return redirect(url_for('student_dashboard'))

    # Fetch questions - IMPORTANT: Know your column order!
    cursor.execute("SELECT id, question, option_a, option_b, option_c, option_d FROM questions WHERE quiz_id = %s ORDER BY id", (quiz_id,))
    questions = cursor.fetchall()
    cursor.close()

    # Debug print to verify structure
    if questions:
        print("Sample question data:", questions[0])
        print("Number of questions:", len(questions))

    if not questions:
        flash("This quiz has no questions yet", "warning")
        return redirect(url_for('student_dashboard'))

    return (render_template('take_quiz.html', quiz=quiz, questions=questions, quiz_id=quiz_id))

@app.route('/submit_quiz/<int:quiz_id>', methods=['POST'])
@login_required()
def submit_quiz(quiz_id):
    cursor = mysql.connection.cursor()
    student_id = session['user_id']

    try:
        # Process submitted answers
        submitted_answers = {}
        for key, value in request.form.items():
            if key.startswith('q'):  # Question answers are named 'q1', 'q2', etc.
                question_id = int(key[1:])  # Extract question ID
                submitted_answers[question_id] = value.upper()  # Store as uppercase (A, B, C, D)

        # Get all questions with correct answers
        cursor.execute("""
            SELECT id, correct_option 
            FROM questions 
            WHERE quiz_id = %s
            ORDER BY id
        """, (quiz_id,))
        questions = cursor.fetchall()
        total_questions = len(questions)

        # Calculate score
        score = 0
        for question in questions:
            question_id = question['id']
            correct_answer = question['correct_option'].upper()
            user_answer = submitted_answers.get(question_id, '')

            if user_answer == correct_answer:
                score += 1

        # Calculate percentage
        percentage = round((score / total_questions) * 100, 2) if total_questions > 0 else 0

        # Save result
        cursor.execute("""
                    INSERT INTO results 
                    (student_id, quiz_id, score, total, taken_at) 
                    VALUES (%s, %s, %s, %s, NOW())
                """, (student_id, quiz_id, score, total_questions))
        mysql.connection.commit()

        # Get quiz details for display
        cursor.execute("SELECT topic FROM quizzes WHERE id = %s", (quiz_id,))
        quiz = cursor.fetchone()

        flash(f"Quiz submitted successfully! Score: {score}/{total_questions} ({percentage}%)", "success")
        return redirect(url_for('quiz_results', quiz_id=quiz_id))

    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error submitting quiz: {str(e)}", "error")
        return redirect(url_for('take_quiz', quiz_id=quiz_id))

    finally:
        cursor.close()
@app.route('/quiz_results/<int:quiz_id>')
@login_required('teacher')
def quiz_results(quiz_id):
    cursor = mysql.connection.cursor()

    # Get quiz details
    cursor.execute("SELECT * FROM quizzes WHERE id = %s", (quiz_id,))
    quiz = cursor.fetchone()

    # Get all results for this quiz
    cursor.execute("""
        SELECT u.username, r.score, r.total, r.taken_at 
        FROM results r
        JOIN users u ON r.student_id = u.id
        WHERE r.quiz_id = %s
        ORDER BY r.score DESC
    """, (quiz_id,))
    results = cursor.fetchall()

    cursor.close()
    return render_template('quiz_result.html', quiz=quiz, results=results)


if __name__ == '__main__':
    app.run(debug=True)
